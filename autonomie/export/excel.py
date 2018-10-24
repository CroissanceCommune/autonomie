# -*- coding: utf-8 -*-
# * Authors:
#       * TJEBBES Gaston <g.t@majerti.fr>
#       * Arezki Feth <f.a@majerti.fr>;
#       * Miotte Julien <j.m@majerti.fr>;
import logging
import openpyxl
import cStringIO
from zope.interface import (
    implements,
)
from openpyxl.styles import (
    Color,
    fills,
    Style,
    NumberFormat,
    PatternFill,
    Font,
)
from autonomie.interfaces import IExporter
from autonomie.export.utils import write_file_to_request

logger = logging.getLogger(__name__)

Color.LightCyan = "FFE0FFFF"
Color.LightCoral = "FFF08080"
Color.LightGreen = "FF90EE90"
Color.Crimson = "FFDC143C"
Color.header = "FFD9EDF7"
Color.footer = "FFFCF8E3"
Color.highlight = "FFEFFFEF"
EXCEL_NUMBER_FORMAT = '0.00'

TITLE_STYLE = Style(font=Font(size=16, bold=True))
HEADER_STYLE = Style(
    font=Font(bold=True),
    fill=PatternFill(
        fill_type=fills.FILL_SOLID,
        start_color=Color(rgb=Color.header)
    )
)
BOLD_CELL = Style(
    font=Font(bold=True)
)
NUMBER_CELL = Style(
    number_format=NumberFormat(format_code=EXCEL_NUMBER_FORMAT)
)
HIGHLIGHTED_ROW_STYLE = Style(
    font=Font(bold=True),
    fill=PatternFill(
        fill_type=fills.FILL_SOLID,
        start_color=Color(rgb=Color.highlight)
    )
)

FOOTER_CELL = Style(
    font=Font(bold=True),
    fill=PatternFill(
        fill_type=fills.FILL_SOLID,
        start_color=Color(rgb=Color.footer)
    ),
    number_format=NumberFormat(format_code=EXCEL_NUMBER_FORMAT)
)
LARGE_FOOTER_CELL = Style(
    font=Font(bold=True, size=16),
    fill=PatternFill(
        fill_type=fills.FILL_SOLID,
        start_color=Color(rgb=Color.footer)
    ),
    number_format=NumberFormat(format_code=EXCEL_NUMBER_FORMAT)
)


class XlsExporter(object):
    implements(IExporter)
    title = u"Export"

    def __init__(self, guess_types=True, worksheet=None, **kw):
        if worksheet is None:
            self.book = openpyxl.workbook.Workbook(guess_types=guess_types)
            self.worksheet = self.book.active
            self.worksheet.title = self.title
        else:
            self.worksheet = worksheet
            self.book = worksheet.parent
        self.options = kw
        self.current_line = 1

    def add_title(self, title, width):
        self.worksheet.merge_cells(
            start_row=self.current_line - 1,
            end_row=self.current_line - 1,
            start_column=0,
            end_column=width - 1,
        )
        cell = self.worksheet.cell(row=self.current_line, column=1)
        cell.value = title
        cell.style = TITLE_STYLE
        row_dim = self.worksheet.row_dimensions.get(self.current_line)
        row_dim.height = 20
        self.current_line += 1

    def add_breakline(self):
        self.current_line += 1

    def _add_row(self, labels, styles=None):
        for col_index, label in enumerate(labels):
            cell = self.worksheet.cell(
                row=self.current_line, column=col_index + 1
            )
            cell.value = label
            if styles:
                cell.style = styles

    def add_headers(self, labels):
        self._add_row(labels, HEADER_STYLE)
        self.current_line += 1

    def add_row(self, labels):
        self._add_row(labels)
        self.current_line += 1

    def add_highlighted_row(self, labels):
        self._add_row(labels, HIGHLIGHTED_ROW_STYLE)
        self.current_line += 1

    def save_book(self, f_buf=None):
        """
        Return a file buffer containing the resulting xls

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        if f_buf is None:
            f_buf = cStringIO.StringIO()
        f_buf.write(openpyxl.writer.excel.save_virtual_workbook(self.book))
        f_buf.seek(0)
        return f_buf

    def render(self, f_buf=None):
        """
        Definitely render the workbook

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        if f_buf is None:
            f_buf = cStringIO.StringIO()

        return self.save_book(f_buf)


def make_excel_view(filename_builder, factory):
    """
        Build an excel view of a model
        :param filename_builder: a callable that take the request as arg and
            return a filename
        :param factory: the Xls factory that should be used to wrap the
            request context the factory should provide a render method
            returning a file like object
    """
    def _view(request):
        """
            the dynamically built view object
        """
        filename = filename_builder(request)
        result = factory(request.context).render()
        request = write_file_to_request(request, filename, result)
        return request.response
    return _view
