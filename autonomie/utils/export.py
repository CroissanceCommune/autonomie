# -*- coding: utf-8 -*-
# * File Name : export.py
#
# * Copyright (C) 2012 Gaston TJEBBES <g.t@majerti.fr>
# * Company : Majerti ( http://www.majerti.fr )
#
#   This software is distributed under GPLV3
#   License: http://www.gnu.org/licenses/gpl-3.0.txt
#
# * Creation Date : 13-03-2013
# * Last Modified :
#
# * Project :
#
"""
    Data export module
"""

import openpyxl
from openpyxl.style import Color, Fill
import cStringIO as StringIO

from autonomie.utils.string import force_ascii
from autonomie.utils.math import integer_to_amount

LETTERS = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I',)

Color.LightCyan = "FFE0FFFF"
Color.LightCoral = "FFF08080"
Color.LightGreen = "FF90EE90"
Color.Crimson = "FFDC143C"
Color.header = "FFD9EDF7"
Color.footer = "FFFCF8E3"


class ExcelExpense(object):
    """
        Wrapper for excel export of an expense object
    """
    expense_columns = [('Date', 'date', None, 'A', 'A'),
                       ('Code analytique', 'code', None, 'B', 'C'),
                       ('Description', 'description', None, 'D', 'G'),
                       ('HT', 'ht', integer_to_amount, 'H', 'H'),
                       ('TVA', 'tva', integer_to_amount, 'I', 'I'),
                       ('Total', 'total', integer_to_amount, 'J', 'J')]

    expensekm_columns = [('Date', 'date', None, 'A', 'A'),
                         ('Code analytique', 'code', None, 'B', 'C'),
                         ('Prestation', 'description', None, 'D', 'F'),
                         ('Départ', 'start', None, 'G', 'G'),
                         ('Arrivée', 'end', None, 'H', 'H'),
                         ('Kms', 'km', integer_to_amount, 'I', 'I'),
                         ('Indemnités', 'total', integer_to_amount, 'J', 'J')]

    def __init__(self, expensesheet):
        self.book = openpyxl.workbook.Workbook()
        self.worksheet = self.book.worksheets[0]
        self.model = expensesheet
        self.index = 1

    def get_merged_cells(self, start, end):
        """
            returned merged cells of the current line index
        """
        cell_gap = '{1}{0}:{2}{0}'.format(self.index, start, end)
        self.worksheet.merge_cells(cell_gap)
        cell = self.worksheet.cell('{1}{0}'.format(self.index, start))
        return cell

    def set_color(self, cell, color):
        cell.style.fill.fill_type = Fill.FILL_SOLID
        cell.style.fill.start_color.index = color

    def write_code(self):
        """
            write the company code in the heade
        """
        code = self.model.company.code_compta
        if not code:
            code = u"Code non renseigné"
        cell = self.get_merged_cells('A', 'D')
        cell.value = u"Code analytique de l'entreprise"
        cell = self.get_merged_cells('E', 'J')
        cell.value = code
        self.index += 1

    def write_user(self):
        """
            write the username in the header
        """
        user = self.model.user
        title = u"%s %s" % (user.lastname, user.firstname)
        cell = self.get_merged_cells('A', 'D')
        cell.value = u"Nom de l'entrepreneur"
        cell = self.get_merged_cells('E', 'J')
        cell.value = title
        self.index += 1

    def write_period(self):
        """
            write the period in the header
        """
        period = "01/{0}/{1}".format(self.model.month,
                self.model.year)
        cell = self.get_merged_cells('A', 'D')
        cell.value = u"Période de la demande"
        cell = self.get_merged_cells('E', 'J')
        cell.value = period
        self.index += 2

    def write_table(self, columns, lines):
        """
            write a table with headers and content
            :param columns: list of (label, model attribute, formatter) where
                formatter could be None
            :params lines: list of models to be written
        """
        for column in columns:
            cell = self.get_merged_cells(column[3], column[4])
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = Color.header
            cell.style.font.bold = True
            cell.value = column[0]
        self.index += 1
        for line in lines:
            for column, letter in zip(columns, LETTERS):
                cell = self.get_merged_cells(column[3], column[4])
                val = getattr(line, column[1])
                if column[2] is not None:
                    val = column[2](val)
                cell.value = val
            self.index += 1

        cell = self.get_merged_cells('A', 'I')
        cell.value = 'Total'
        cell.style.font.bold = True
        self.set_color(cell, Color.footer)
        cell = self.get_merged_cells('J', 'J')
        cell.style.font.bold = True
        cell.value = integer_to_amount(sum([line.total for line in lines]))
        self.set_color(cell, Color.footer)
        self.index += 2

    def write_expense_table(self, category):
        """
            write expenses tables for the given category
        """
        lines = [line for line in self.model.lines
                            if line.category == category]
        columns = self.expense_columns
        self.write_table(columns, lines)
        kmlines = [lin for lin in self.model.kmlines
                            if lin.category == category]
        columns = self.expensekm_columns
        self.write_table(columns, kmlines)

    def write_full_line(self, txt, start="A", end="J"):
        """
            Write a full line, merging cells
        """
        cell = self.get_merged_cells(start, end)
        cell.value = txt
        self.index += 1
        return cell

    def write_internal_expenses(self):
        """
            write the internal expense table to the current worksheet
        """
        txt = u"FRAIS DIRECT DE FONCTIONNEMENT (< à 30% DU SALAIRE BRUT \
PAR MOIS)"
        cell = self.write_full_line(txt)
        cell.style.font.color.index = Color.Crimson
        self.write_expense_table('1')

    def write_activity_expenses(self):
        """

            write the activity expense table to the current worksheet
        """
        txt = u"FRAIS CONCERNANT DIRECTEMENT VOTRE L'ACTIVITE AUPRES DE VOS \
CLIENTS"
        cell = self.write_full_line(txt)
        cell.style.font.color.index = Color.Crimson
        self.write_expense_table('2')

    def write_total(self):
        """
            write the final total
        """
        cell = self.get_merged_cells('D', 'I')
        cell.value = u"Total des frais professionnel à payer"
        cell.style.font.bold = True
        cell.style.font.size = 16
        self.set_color(cell, Color.footer)
        cell = self.get_merged_cells('J', 'J')
        cell.style.font.bold = True
        cell.style.font.size = 16
        cell.value = integer_to_amount(self.model.total)
        self.set_color(cell, Color.footer)
        self.index += 2

    def write_accord(self):
        """
            Write the endline
        """
        cell = self.get_merged_cells('D', 'J')
        cell.value = u"Accord après vérification"

    def render(self):
        """
            Return the current excel export as a String buffer (StringIO)
        """
        cell = self.write_full_line(u"Feuille de notes de frais")
        cell.style.font.bold = True
        cell.style.font.size = 24
        self.index += 2

        self.write_code()
        self.write_user()
        self.write_period()
        self.write_internal_expenses()
        self.write_activity_expenses()
        self.write_total()
        result = StringIO.StringIO()
        self.book.save(result)
        return result


def write_excel_headers(request, filename):
    """
        write the headers of the excel file 'filename'
    """
    request.response.content_type = 'application/vnd.ms-excel'
    request.response.headerlist.append(
            ('Content-Disposition',
                'attachment; filename={0}'.format(force_ascii(filename))))
    return request


def write_excel(request, filename, factory):
    """
        write an excel stylesheet to the current request
        :param filename: the filename to output the document to
        :param factory: the Excel factory that should be used to wrap the
            request context the factory should provide a render method
            returning a file like object
    """
    request = write_excel_headers(request, filename)
    result = factory(request.context).render()
    request.response.write(result.getvalue())
    return request


def make_excel_view(filename_builder, factory):
    """
        Build an excel view of a model
        :param filename_builder: a callable that take the request as arg and
            return a filename
        :param factory: the Excel factory that should be used to wrap the
            request context the factory should provide a render method
            returning a file like object
    """
    def _view(request):
        """
            the dynamically builded view object
        """
        filename = filename_builder(request)
        request = write_excel(request, filename, factory)
        return request.response
    return _view
